"""
Content calendar enrichment tool.

Reads a content calendar Excel file, enriches each row across three platforms:
- TikTok: trends, audio picks, and remix ideas
- Twitter: draft tweets, thread structures, and references
- LinkedIn: draft posts and references

Outputs enriched Excel (3 sheets) and diagnostic run log.

Usage: python enrich_calendar.py --input "Content ideas.xlsx"
"""
import argparse
import datetime
import math
import os
import sys
import time
import json
from collections import Counter
from dataclasses import dataclass
from typing import Tuple, List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from dateutil.parser import parse as date_parse
import requests
import openai
from dotenv import load_dotenv

load_dotenv()


VALID_TYPES = {"Story", "BTS", "Teach", "Trend", "Breakdown", "Reflection", "Depth"}

# RapidAPI Configuration
RAPIDAPI_TIKTOK_HOST = "tiktok-api23.p.rapidapi.com"
RAPIDAPI_TWITTER_HOST = "twitter-api45.p.rapidapi.com"
MAX_RESULTS_PER_ROW = 21
RESULTS_PER_QUERY = 10
MIN_LIKES = 1_000
MAX_RETRIES = 3

# Twitter Configuration
TWITTER_RESULTS_PER_QUERY = 10
TWITTER_MAX_RESULTS_PER_ROW = 20

# Query caches to avoid duplicate API calls
_query_cache: Dict[str, List[Dict]] = {}
_twitter_query_cache: Dict[str, List[Dict]] = {}


def extract_keywords(text: str) -> List[str]:
    """Extract meaningful keywords (>3 chars, no stop words) from text."""
    import re
    # Strip punctuation from each word before filtering
    words = [re.sub(r'[^\w]', '', w) for w in text.lower().split()]
    stop_words = {'the', 'and', 'for', 'with', 'from', 'this', 'that', 'what',
                  'when', 'where', 'who', 'how', 'why'}
    return [w for w in words if len(w) > 3 and w not in stop_words]


def has_enough_context(idea) -> bool:
    """Check if a row has enough context to justify an API call (>= 3 keywords)."""
    all_kw = extract_keywords(idea.topic) + extract_keywords(idea.description)
    return len(all_kw) >= 3


@dataclass
class ContentIdea:
    """Represents a single content idea from the calendar."""
    row_number: int
    date: datetime.date
    content_type: str
    topic: str
    description: str
    idea_text: str


def normalize_text(text: str) -> str:
    """
    Normalize text by replacing special Unicode characters and cleaning whitespace.

    Args:
        text: Input text to normalize

    Returns:
        Normalized text with regular hyphens/spaces and collapsed whitespace
    """
    if not text:
        return ""

    # Replace non-breaking hyphen (U+2011) with regular hyphen
    text = text.replace('\u2011', '-')

    # Replace non-breaking space (U+00A0) with regular space
    text = text.replace('\u00a0', ' ')

    # Strip leading/trailing whitespace
    text = text.strip()

    # Collapse multiple internal spaces into single space
    text = ' '.join(text.split())

    return text


def sentence_case(text: str) -> str:
    """
    Convert text to sentence case (capitalize first letter, preserve rest).

    Args:
        text: Input text

    Returns:
        Text with first character capitalized, rest preserved as-is
    """
    if not text:
        return ""

    return text[0].upper() + text[1:] if len(text) > 1 else text.upper()


def parse_date(date_str: str) -> datetime.date:
    """
    Parse date string into date object, handling special Unicode characters.

    Args:
        date_str: Date string to parse

    Returns:
        Parsed date object

    Raises:
        ValueError: If date cannot be parsed
    """
    # First normalize to handle special hyphens
    normalized = normalize_text(date_str)

    try:
        return date_parse(normalized).date()
    except Exception as e:
        raise ValueError(f"Failed to parse date '{date_str}': {e}")


def load_content_ideas(file_path: str) -> Tuple[List[ContentIdea], List[Dict[str, Any]]]:
    """
    Load and validate content ideas from Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        Tuple of (list of ContentIdea objects, list of skipped row info dicts)

    Raises:
        ValueError: If required columns are missing or validation fails
    """
    ideas = []
    skipped = []

    # Open workbook
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)

    # Try to get Sheet1, fall back to first sheet
    if "Sheet1" in wb.sheetnames:
        sheet = wb["Sheet1"]
        print(f"Reading from sheet: Sheet1")
    else:
        sheet = wb.active
        print(f"Sheet1 not found, using first sheet: {sheet.title}")

    # Read and validate header row
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

    # Create case-insensitive column mapping
    header_map = {col.lower().strip() if col else "": idx
                  for idx, col in enumerate(header_row)}

    required_columns = ["day", "date", "type", "topic", "description"]
    missing_columns = []

    for col in required_columns:
        if col not in header_map:
            missing_columns.append(col)

    if missing_columns:
        found_columns = [col for col in header_row if col]
        raise ValueError(
            f"Missing required columns: {missing_columns}\n"
            f"Expected: {required_columns}\n"
            f"Found: {found_columns}"
        )

    # Map column positions
    col_date = header_map["date"]
    col_type = header_map["type"]
    col_topic = header_map["topic"]
    col_description = header_map["description"]

    # Process data rows (row 2+)
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract values
        date_val = row[col_date] if col_date < len(row) else None
        type_val = row[col_type] if col_type < len(row) else None
        topic_val = row[col_topic] if col_topic < len(row) else None
        desc_val = row[col_description] if col_description < len(row) else None

        # Normalize all text values
        date_str = normalize_text(str(date_val)) if date_val is not None else ""
        type_str = normalize_text(str(type_val)) if type_val is not None else ""
        topic_str = normalize_text(str(topic_val)) if topic_val is not None else ""
        desc_str = normalize_text(str(desc_val)) if desc_val is not None else ""

        # Check for missing required fields
        if not type_str:
            skipped.append({"row_number": row_idx, "field": "Type"})
            continue
        if not topic_str:
            skipped.append({"row_number": row_idx, "field": "Topic"})
            continue
        if not desc_str:
            skipped.append({"row_number": row_idx, "field": "Description"})
            continue
        if not date_str:
            skipped.append({"row_number": row_idx, "field": "Date"})
            continue

        # Validate Type (case-insensitive) and find proper casing
        type_lower = type_str.lower()
        valid_types_map = {t.lower(): t for t in VALID_TYPES}

        if type_lower not in valid_types_map:
            raise ValueError(
                f"Invalid Type '{type_str}' at row {row_idx}\n"
                f"Valid types: {sorted(VALID_TYPES)}"
            )

        # Use the properly-cased type from VALID_TYPES
        content_type = valid_types_map[type_lower]

        # Sentence-case Topic and Description
        topic = sentence_case(topic_str)
        description = sentence_case(desc_str)

        # Parse date
        try:
            date_obj = parse_date(date_str)
        except ValueError as e:
            raise ValueError(f"Row {row_idx}: {e}")

        # Construct idea_text
        idea_text = f"{content_type} | {topic} | {description}"

        # Create ContentIdea
        idea = ContentIdea(
            row_number=row_idx,
            date=date_obj,
            content_type=content_type,
            topic=topic,
            description=description,
            idea_text=idea_text
        )
        ideas.append(idea)

    wb.close()

    return ideas, skipped


def print_summary(ideas: List[ContentIdea], skipped: List[Dict[str, Any]]) -> None:
    """
    Print summary of loaded content ideas.

    Args:
        ideas: List of loaded ContentIdea objects
        skipped: List of skipped row information
    """
    print(f"Loaded {len(ideas)} content ideas")

    # Get unique types
    unique_types = sorted(set(idea.content_type for idea in ideas))
    print(f"Types: {', '.join(unique_types)}")

    # Get date range
    if ideas:
        dates = [idea.date for idea in ideas]
        earliest = min(dates)
        latest = max(dates)
        print(f"Date range: {earliest} to {latest}")

    # Print skipped rows
    if skipped:
        print("\nSkipped rows:")
        for skip_info in skipped:
            print(f"  Skipped row {skip_info['row_number']}: missing {skip_info['field']}")


# Query generation for TikTok search
TYPE_QUERY_MAP = {
    "Story": "entrepreneur story time",
    "BTS": "day in the life CEO",
    "Teach": "business lessons learned",
    "Trend": "entrepreneur new chapter",
    "Breakdown": "business mindset shift",
    "Reflection": "entrepreneur honest reflection",
    "Depth": "founder weekly vlog"
}


def generate_queries(idea: ContentIdea) -> List[str]:
    """
    Generate up to 3 targeted TikTok search queries from a content idea.

    Priority: format-specific > topic > keyword combo.
    All queries are <=5 words.
    """
    queries = []

    def truncate_to_5_words(text: str) -> str:
        words = text.split()
        return ' '.join(words[:5]) if len(words) > 5 else text

    # 1. Format-specific query (highest value)
    queries.append(TYPE_QUERY_MAP.get(idea.content_type, "founder content"))

    # 2. Topic as query
    if len(idea.topic.split()) <= 5:
        queries.append(idea.topic.lower())

    # 3. Best keyword combo from topic + description
    all_kw = extract_keywords(idea.topic) + extract_keywords(idea.description)
    if len(all_kw) >= 2:
        queries.append(truncate_to_5_words(' '.join(all_kw[:3])))

    # Deduplicate
    seen = set()
    unique = []
    for q in queries:
        q_norm = q.lower().strip()
        if q_norm not in seen:
            seen.add(q_norm)
            unique.append(q_norm)

    # Ensure minimum 2 with fallback
    if len(unique) < 2:
        unique.append("startup founder content")

    return unique[:3]


def _call_with_retry(fn, *args, max_retries=MAX_RETRIES, **kwargs) -> Any:
    """
    Call function with exponential backoff retry on failures.

    Args:
        fn: Function to call
        *args: Positional arguments for fn
        max_retries: Maximum number of retry attempts
        **kwargs: Keyword arguments for fn

    Returns:
        Result from fn, or empty list on final failure
    """
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except (requests.RequestException, ValueError, openai.APIError) as e:
            # Don't retry billing/auth errors — they'll never succeed
            err_str = str(e)
            if any(code in err_str for code in ("401", "402", "403", "insufficient_quota", "hard limit")):
                print(f"Non-retryable error: {e}", file=sys.stderr)
                return []
            if attempt < max_retries - 1:
                wait_time = 2 ** attempt
                print(f"Retry {attempt + 1}/{max_retries} after error: {e}", file=sys.stderr)
                print(f"Waiting {wait_time}s before retry...", file=sys.stderr)
                time.sleep(wait_time)
            else:
                print(f"Failed after {max_retries} attempts: {e}", file=sys.stderr)
                return []


def _run_tiktok_rapidapi(query: str, api_key: str, max_items: int = RESULTS_PER_QUERY) -> List[Dict[str, Any]]:
    """Fetch TikTok search results via RapidAPI tiktok-api23."""
    url = f"https://{RAPIDAPI_TIKTOK_HOST}/api/search/general"
    headers = {
        "x-rapidapi-host": RAPIDAPI_TIKTOK_HOST,
        "x-rapidapi-key": api_key,
    }
    params = {"keyword": query, "count": max_items}

    response = requests.get(url, headers=headers, params=params, timeout=60)

    if response.status_code != 200:
        raise requests.RequestException(
            f"TikTok API returned status {response.status_code}: {response.text}"
        )

    data = response.json()
    # Extract item dicts from data array
    results = []
    for entry in data.get("data", []):
        item = entry.get("item")
        if item:
            results.append(item)
    return results[:max_items]


def fetch_tiktok_results(queries: List[str], token: Optional[str] = None) -> List[Dict[str, Any]]:
    """Fetch TikTok results for multiple queries with caching and result cap."""
    if token is None:
        token = os.environ.get("RAPIDAPI_KEY")
    if not token:
        raise ValueError("RAPIDAPI_KEY not found in environment")

    results = []
    for query in queries:
        query_normalized = query.lower().strip()

        if query_normalized in _query_cache:
            print(f"TikTok cache hit: {query}", file=sys.stderr)
            results.extend(_query_cache[query_normalized])
        else:
            query_results = _call_with_retry(_run_tiktok_rapidapi, query, token)
            _query_cache[query_normalized] = query_results
            results.extend(query_results)

        if len(results) >= MAX_RESULTS_PER_ROW:
            results = results[:MAX_RESULTS_PER_ROW]
            break

    return results


def enrich_row_queries(idea: ContentIdea) -> Dict[str, Any]:
    """
    Generate queries for a content idea (Phase 2 pipeline helper).

    This structural helper wires query generation to row context.
    Phase 3+ will extend this to include TikTok results and analysis.

    Args:
        idea: ContentIdea to generate queries for

    Returns:
        Dict with row_number, queries list, and query_count
    """
    queries = generate_queries(idea)

    return {
        "row_number": idea.row_number,
        "queries": queries,
        "query_count": len(queries)
    }


def normalize_result(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Normalize RapidAPI TikTok result to standard schema.

    Maps tiktok-api23 item schema to our standard video metadata structure.
    Returns None for malformed results (missing required fields).
    """
    video_id = raw.get("id")
    if not video_id:
        return None

    stats = raw.get("stats", {})
    author = raw.get("author", {})
    music = raw.get("music", {})

    # Convert unix timestamp to ISO string
    create_time = raw.get("createTime")
    create_time_iso = None
    if create_time:
        try:
            create_time_iso = datetime.datetime.fromtimestamp(
                int(create_time), tz=datetime.timezone.utc
            ).isoformat()
        except (ValueError, OSError):
            pass

    return {
        "video_id": str(video_id),
        "url": f"https://www.tiktok.com/@{author.get('uniqueId', '')}/video/{video_id}",
        "caption": raw.get("desc", ""),
        "author_username": author.get("uniqueId", ""),
        "create_time": create_time_iso,
        "views": stats.get("playCount", 0),
        "likes": stats.get("diggCount", 0),
        "comments": stats.get("commentCount", 0),
        "shares": stats.get("shareCount", 0),
        "audio": {
            "audio_id": str(music.get("id", "")),
            "title": music.get("title", ""),
            "author": music.get("authorName", ""),
            "url": music.get("playUrl", "")
        }
    }


def deduplicate_results(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Remove duplicate videos by video_id, keeping first occurrence.

    Args:
        results: List of normalized video dicts

    Returns:
        Deduplicated list preserving original order
    """
    seen = set()
    deduped = []

    for result in results:
        video_id = result.get("video_id")
        if video_id and video_id not in seen:
            seen.add(video_id)
            deduped.append(result)

    return deduped


def filter_old_results(results: List[Dict[str, Any]], max_age_days: int = 120) -> List[Dict[str, Any]]:
    """
    Filter out videos older than max_age_days.

    Videos with missing or unparseable create_time are KEPT (no false rejection).

    Args:
        results: List of normalized video dicts
        max_age_days: Maximum age in days (default 120)

    Returns:
        Filtered list with only recent videos
    """
    filtered = []
    now = datetime.datetime.now(datetime.timezone.utc)

    for result in results:
        create_time = result.get("create_time")

        # Keep if create_time is missing or empty
        if not create_time:
            filtered.append(result)
            continue

        # Try to parse and check age
        try:
            created_dt = date_parse(create_time)
            # Ensure timezone-aware
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=datetime.timezone.utc)

            age_days = (now - created_dt).days

            if age_days <= max_age_days:
                filtered.append(result)
        except Exception:
            # Keep on parsing failure (graceful fallback)
            filtered.append(result)

    return filtered


def score_video(video: Dict[str, Any], keywords: List[str] = None) -> Dict[str, Any]:
    """
    Compute engagement score for a video.

    Scoring formula (DAT-04 to DAT-07):
    - Base score: log10(views+1)*0.65 + log10(likes+1)*0.25 + log10(comments+1)*0.10
    - Recency boost: +0.2 for <=14 days, +0.1 for <=30 days, 0 otherwise
    - Relevance boost: +0.05 per keyword in caption, capped at +0.2
    - Final score: base + recency_boost + relevance_boost

    Args:
        video: Normalized video dict
        keywords: Optional list of keywords for relevance scoring

    Returns:
        Video dict with added score fields (base_score, recency_boost, relevance_boost, final_score)
    """
    # Step 1: Base score (DAT-04)
    views = video.get("views", 0) or 0
    likes = video.get("likes", 0) or 0
    comments = video.get("comments", 0) or 0

    # Like-to-view ratio bonus (rewards genuine engagement, penalizes AI slop)
    like_ratio = likes / views if views > 0 else 0
    # Normalize: 3%+ ratio = full bonus (1.0), scale linearly below that
    ratio_score = min(like_ratio / 0.03, 1.0)

    base = (math.log10(likes + 1) * 0.45 +
            math.log10(comments + 1) * 0.25 +
            math.log10(views + 1) * 0.15 +
            ratio_score * 0.15)

    # Step 2: Recency boost (DAT-05)
    recency_boost = 0.0
    create_time = video.get("create_time")

    if create_time:
        try:
            created_dt = date_parse(create_time)
            # Ensure timezone-aware
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=datetime.timezone.utc)

            now = datetime.datetime.now(datetime.timezone.utc)
            days_old = (now - created_dt).days

            if days_old <= 14:
                recency_boost = 0.2
            elif days_old <= 30:
                recency_boost = 0.1
        except Exception:
            # On parsing failure, no recency boost
            pass

    # Step 3: Relevance boost (DAT-06)
    relevance_boost = 0.0

    if keywords:
        caption = video.get("caption", "")
        if caption:
            caption_lower = caption.lower()
            overlap_count = sum(1 for kw in keywords if kw.lower() in caption_lower)
            relevance_boost = min(overlap_count * 0.05, 0.2)

    # Step 4: Final score (DAT-07)
    final_score = base + recency_boost + relevance_boost

    # Add scores to video dict
    video["base_score"] = base
    video["recency_boost"] = recency_boost
    video["relevance_boost"] = relevance_boost
    video["final_score"] = final_score

    return video


def process_results(raw_results: List[Dict[str, Any]], keywords: List[str] = None) -> List[Dict[str, Any]]:
    """
    Process raw Apify results through full pipeline.

    Pipeline:
    1. Normalize: Convert Apify schema to standard schema
    2. Deduplicate: Remove duplicate video_ids
    3. Filter: Remove videos older than 120 days
    4. Score: Compute engagement scores with optional keyword relevance
    5. Sort: By final_score descending

    Args:
        raw_results: List of raw Apify result dicts
        keywords: Optional list of keywords for relevance scoring

    Returns:
        Sorted list of scored, normalized video dicts
    """
    # Step 1: Normalize (filter out None results)
    normalized = [normalize_result(r) for r in raw_results]
    normalized = [n for n in normalized if n is not None]

    # Step 2: Deduplicate
    deduped = deduplicate_results(normalized)

    # Step 3: Filter old
    filtered = filter_old_results(deduped)

    # Step 4: Score
    scored = [score_video(v, keywords) for v in filtered]

    # Step 5: Sort by final_score descending
    sorted_results = sorted(scored, key=lambda x: x.get("final_score", 0), reverse=True)

    return sorted_results


def select_top_examples(scored_results: List[Dict[str, Any]], count: int = 3) -> List[Dict[str, Any]]:
    """
    Select top N videos from scored results as examples.

    Takes already-sorted (descending by final_score) output from process_results
    and extracts top N videos with relevant metadata for content creation.

    Args:
        scored_results: Sorted list of scored video dicts from process_results()
        count: Number of top examples to select (default 3)

    Returns:
        List of up to 'count' example dicts with url, engagement metrics,
        author info, caption, audio, and final_score. Returns fewer if less available.
        Returns empty list if scored_results is empty.
    """
    if not scored_results:
        return []

    examples = []
    for video in scored_results[:count]:
        example = {
            "url": video["url"],
            "views": video["views"],
            "likes": video["likes"],
            "comments": video["comments"],
            "shares": video["shares"],
            "author_username": video["author_username"],
            "create_time": video["create_time"],
            "caption": video["caption"],
            "audio_title": video.get("audio", {}).get("title", ""),
            "final_score": video["final_score"]
        }
        examples.append(example)

    return examples


def select_audio(scored_results: List[Dict[str, Any]], examples: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Select recommended audio track from top scored results.

    Strategy:
    1. Count audio_id occurrences among top 20 results
    2. Select most common audio_id (if any repeats found)
    3. Fall back to first example's audio if no repeats
    4. Determine confidence based on occurrence count

    Args:
        scored_results: Sorted list of scored video dicts from process_results()
        examples: List of example dicts from select_top_examples()

    Returns:
        Dict with audio_title, audio_author, audio_id, audio_url, audio_confidence.
        Confidence levels: "high" (>=3 occurrences), "medium" (2), "low" (1 or fallback).
        Returns empty strings with "low" confidence if no audio found.
    """
    # Step 1: Count audio occurrences in top 20
    top_20 = scored_results[:20]
    audio_ids = []

    for video in top_20:
        audio_id = video.get("audio", {}).get("audio_id", "")
        if audio_id:  # Skip empty/falsy audio_ids
            audio_ids.append(audio_id)

    audio_counter = Counter(audio_ids)

    # Step 2: Select audio
    selected_audio_id = None
    count = 0

    if audio_counter:
        # Get most common audio_id
        selected_audio_id, count = audio_counter.most_common(1)[0]

        # Find first video with this audio_id to get full metadata
        selected_video = None
        for video in top_20:
            if video.get("audio", {}).get("audio_id") == selected_audio_id:
                selected_video = video
                break

        if selected_video:
            audio_title = selected_video.get("audio", {}).get("title", "")
            audio_author = selected_video.get("audio", {}).get("author", "")
            audio_url = selected_video.get("audio", {}).get("url", "")
        else:
            # Shouldn't happen, but fallback
            audio_title = ""
            audio_author = ""
            audio_url = ""
    else:
        # No valid audio_ids in top 20, fall back to example 1
        if examples and len(examples) > 0:
            # Find the video in scored_results that matches example 1 URL
            ex1_url = examples[0].get("url", "")
            selected_video = None
            for video in scored_results:
                if video.get("url") == ex1_url:
                    selected_video = video
                    break

            if selected_video:
                selected_audio_id = selected_video.get("audio", {}).get("audio_id", "")
                audio_title = selected_video.get("audio", {}).get("title", "")
                audio_author = selected_video.get("audio", {}).get("author", "")
                audio_url = selected_video.get("audio", {}).get("url", "")
                count = 1  # Fallback used
            else:
                # Can't find ex1 video
                selected_audio_id = ""
                audio_title = ""
                audio_author = ""
                audio_url = ""
                count = 0
        else:
            # No examples either
            selected_audio_id = ""
            audio_title = ""
            audio_author = ""
            audio_url = ""
            count = 0

    # Step 3: Determine confidence
    if count >= 3:
        confidence = "high"
    elif count == 2:
        confidence = "medium"
    else:
        confidence = "low"

    # Step 4: Build return dict
    return {
        "audio_title": audio_title,
        "audio_author": audio_author,
        "audio_id": selected_audio_id,
        "audio_url": audio_url,
        "audio_confidence": confidence
    }


# ---------------------------------------------------------------------------
# Twitter integration
# ---------------------------------------------------------------------------

TWITTER_TYPE_QUERY_MAP = {
    "Story": "my story entrepreneurship",
    "BTS": "building in public",
    "Teach": "business advice",
    "Trend": "entrepreneur hot take",
    "Breakdown": "business breakdown",
    "Reflection": "founder lessons",
    "Depth": "entrepreneurship thread"
}


def generate_twitter_queries(idea: ContentIdea) -> List[str]:
    """Generate up to 3 Twitter search queries (hashtag + topic style)."""
    queries = []

    def truncate_to_5_words(text: str) -> str:
        words = text.split()
        return ' '.join(words[:5]) if len(words) > 5 else text

    # 1. Format-specific hashtag query (highest value)
    queries.append(TWITTER_TYPE_QUERY_MAP.get(idea.content_type, "#startup founder"))

    # 2. Topic as query
    if len(idea.topic.split()) <= 5:
        queries.append(idea.topic.lower())

    # 3. Best keyword combo
    all_kw = extract_keywords(idea.topic) + extract_keywords(idea.description)
    if len(all_kw) >= 2:
        queries.append(truncate_to_5_words(' '.join(all_kw[:3])))

    # Deduplicate
    seen = set()
    unique = []
    for q in queries:
        q_norm = q.lower().strip()
        if q_norm not in seen:
            seen.add(q_norm)
            unique.append(q_norm)

    if len(unique) < 2:
        unique.append("startup founder advice")

    return unique[:3]


def _run_twitter_rapidapi(query: str, api_key: str, max_items: int = TWITTER_RESULTS_PER_QUERY) -> List[Dict[str, Any]]:
    """Fetch Twitter search results via RapidAPI twitter-api45."""
    url = f"https://{RAPIDAPI_TWITTER_HOST}/search.php"
    headers = {
        "x-rapidapi-host": RAPIDAPI_TWITTER_HOST,
        "x-rapidapi-key": api_key,
    }
    params = {"query": query, "search_type": "Top"}
    response = requests.get(url, headers=headers, params=params, timeout=60)

    if response.status_code != 200:
        raise requests.RequestException(
            f"Twitter API returned status {response.status_code}: {response.text}"
        )

    data = response.json()
    results = data.get("timeline", [])
    return results[:max_items]


def fetch_twitter_results(queries: List[str], token: Optional[str] = None) -> List[Dict[str, Any]]:
    """Fetch Twitter results for multiple queries with caching and result cap."""
    if token is None:
        token = os.environ.get("RAPIDAPI_KEY")
    if not token:
        raise ValueError("RAPIDAPI_KEY not found in environment")

    results = []
    for query in queries:
        query_normalized = query.lower().strip()
        if query_normalized in _twitter_query_cache:
            print(f"Twitter cache hit: {query}", file=sys.stderr)
            results.extend(_twitter_query_cache[query_normalized])
        else:
            query_results = _call_with_retry(_run_twitter_rapidapi, query, token)
            _twitter_query_cache[query_normalized] = query_results
            results.extend(query_results)

        if len(results) >= TWITTER_MAX_RESULTS_PER_ROW:
            results = results[:TWITTER_MAX_RESULTS_PER_ROW]
            break

    return results


def normalize_twitter_result(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Normalize RapidAPI twitter-api45 tweet result to standard schema."""
    tweet_id = raw.get("tweet_id") or raw.get("id")
    if not tweet_id:
        return None

    screen_name = raw.get("screen_name", "")

    # views comes as string from this API
    views_raw = raw.get("views", 0)
    try:
        views = int(views_raw) if views_raw else 0
    except (ValueError, TypeError):
        views = 0

    return {
        "tweet_id": str(tweet_id),
        "url": f"https://twitter.com/{screen_name}/status/{tweet_id}" if screen_name else f"https://twitter.com/i/status/{tweet_id}",
        "text": raw.get("text", ""),
        "author": screen_name,
        "likes": raw.get("favorites", 0) or 0,
        "retweets": raw.get("retweets", 0) or 0,
        "replies": raw.get("replies", 0) or 0,
        "views": views,
        "created_at": raw.get("created_at", None),
    }


def score_tweet(tweet: Dict[str, Any], keywords: List[str] = None) -> Dict[str, Any]:
    """Score a tweet: log10(likes+1)*0.45 + log10(retweets+1)*0.30 + log10(replies+1)*0.25 + boosts."""
    likes = tweet.get("likes", 0) or 0
    retweets = tweet.get("retweets", 0) or 0
    replies = tweet.get("replies", 0) or 0

    base = (math.log10(likes + 1) * 0.45 +
            math.log10(retweets + 1) * 0.30 +
            math.log10(replies + 1) * 0.25)

    # Recency boost
    recency_boost = 0.0
    created_at = tweet.get("created_at")
    if created_at:
        try:
            created_dt = date_parse(created_at)
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=datetime.timezone.utc)
            days_old = (datetime.datetime.now(datetime.timezone.utc) - created_dt).days
            if days_old <= 14:
                recency_boost = 0.2
            elif days_old <= 30:
                recency_boost = 0.1
        except Exception:
            pass

    # Relevance boost
    relevance_boost = 0.0
    if keywords:
        text = tweet.get("text", "")
        if text:
            text_lower = text.lower()
            overlap = sum(1 for kw in keywords if kw.lower() in text_lower)
            relevance_boost = min(overlap * 0.05, 0.2)

    tweet["base_score"] = base
    tweet["recency_boost"] = recency_boost
    tweet["relevance_boost"] = relevance_boost
    tweet["final_score"] = base + recency_boost + relevance_boost
    return tweet


def process_twitter_results(raw_results: List[Dict[str, Any]], keywords: List[str] = None) -> List[Dict[str, Any]]:
    """Normalize → dedup → filter old → score → sort tweets."""
    normalized = [normalize_twitter_result(r) for r in raw_results]
    normalized = [n for n in normalized if n is not None]

    # Dedup by tweet_id
    seen = set()
    deduped = []
    for t in normalized:
        tid = t.get("tweet_id")
        if tid and tid not in seen:
            seen.add(tid)
            deduped.append(t)

    # Filter old (120 days)
    filtered = filter_old_results(deduped, max_age_days=120)

    scored = [score_tweet(t, keywords) for t in filtered]
    return sorted(scored, key=lambda x: x.get("final_score", 0), reverse=True)


def select_top_tweets(scored: List[Dict[str, Any]], count: int = 3) -> List[Dict[str, Any]]:
    """Select top N tweets as examples."""
    if not scored:
        return []
    examples = []
    for tweet in scored[:count]:
        examples.append({
            "url": tweet["url"],
            "text": tweet["text"],
            "author": tweet["author"],
            "likes": tweet["likes"],
            "retweets": tweet["retweets"],
            "replies": tweet["replies"],
            "views": tweet["views"],
            "created_at": tweet["created_at"],
            "final_score": tweet["final_score"],
        })
    return examples


def generate_twitter_content(idea: ContentIdea, examples: List[Dict[str, Any]], client=None) -> Dict[str, Any]:
    """Generate Twitter content via LLM: draft_tweet, thread_structure, reference_notes, remix_ideas."""
    empty = {"draft_tweet": "", "thread_structure": "", "reference_notes": "", "remix_ideas": ""}

    if client is None:
        if not os.environ.get("OPENAI_API_KEY"):
            return empty
        client = openai.OpenAI()

    system_message = (
        "You are a Twitter content strategist for Logara, a startup documenting its founder journey. "
        "You craft viral tweets and threads. Respond in valid JSON only."
    )

    examples_text = ""
    for i, ex in enumerate(examples, 1):
        examples_text += f"\nExample {i}:\n"
        examples_text += f"  - Text: {ex.get('text', 'N/A')[:280]}\n"
        examples_text += f"  - Likes: {ex.get('likes', 0):,} | Retweets: {ex.get('retweets', 0):,}\n"
        examples_text += f"  - Author: @{ex.get('author', 'unknown')}\n"

    user_message = f"""Content Idea:
- Type: {idea.content_type}
- Topic: {idea.topic}
- Description: {idea.description}

Top Example Tweets:{examples_text}

Brand Context:
- Logara is a startup documenting its founder journey
- Tone: Authentic, sharp, conversational

Instructions:
1. draft_tweet: A single tweet (max 280 chars) with a compelling hook for this topic
2. thread_structure: A 3-5 tweet thread outline (numbered, each tweet summarized in one line)
3. reference_notes: For each example, one sentence on why it works and what to learn
4. remix_ideas: 2-3 tweet angle ideas as one-liner bullets

Return JSON:
{{
  "draft_tweet": "the tweet text (max 280 chars)",
  "thread_structure": "1. First tweet...\\n2. Second tweet...\\n3. ...",
  "reference_notes": ["note for ex1", "note for ex2", "note for ex3"],
  "remix_ideas": ["angle 1", "angle 2", "angle 3"]
}}"""

    def make_api_call():
        return client.chat.completions.create(
            model="gpt-4o-mini",
            response_format={"type": "json_object"},
            temperature=0.8,
            max_tokens=600,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ]
        )

    try:
        response = _call_with_retry(make_api_call)
        if not response or response == []:
            return empty

        parsed = json.loads(response.choices[0].message.content)
        ref_notes = parsed.get("reference_notes", [])
        remix_list = parsed.get("remix_ideas", [])

        return {
            "draft_tweet": parsed.get("draft_tweet", ""),
            "thread_structure": parsed.get("thread_structure", ""),
            "reference_notes": "\n".join(f"- {n}" for n in ref_notes) if ref_notes else "",
            "remix_ideas": "\n".join(f"- {r}" for r in remix_list) if remix_list else "",
        }
    except (json.JSONDecodeError, KeyError, IndexError, AttributeError) as e:
        print(f"Error parsing Twitter LLM response: {e}", file=sys.stderr)
        return empty


def enrich_row_twitter(idea: ContentIdea, raw_results: List[Dict[str, Any]], openai_client=None) -> Dict[str, Any]:
    """Orchestrate full Twitter enrichment pipeline for a single row."""
    keywords = extract_keywords(idea.topic) + extract_keywords(idea.description)
    scored = process_twitter_results(raw_results, keywords=keywords)
    examples = select_top_tweets(scored)

    enrich_status = "ok"
    enrich_reason = ""

    if len(scored) == 0:
        # No API results — generate content from idea alone (LLM-only fallback)
        llm = generate_twitter_content(idea, [], client=openai_client)
        llm_has_content = any(llm.get(k, "") for k in ["draft_tweet", "thread_structure", "remix_ideas"])
        if llm_has_content:
            enrich_status = "partial"
            enrich_reason = "LLM-only (no API results)"
        else:
            enrich_status = "skipped"
            enrich_reason = "No API results and LLM generation failed"
    else:
        llm = generate_twitter_content(idea, examples, client=openai_client)
        llm_failed = all(llm.get(k, "") == "" for k in ["draft_tweet", "thread_structure", "reference_notes", "remix_ideas"])
        reasons = []
        if llm_failed:
            reasons.append("LLM generation failed")
        if len(examples) < 2:
            reasons.append(f"Only {len(examples)} example(s), need >=2")
        if reasons:
            enrich_status = "partial"
            enrich_reason = "; ".join(reasons)

    return {
        "row_number": idea.row_number,
        "content_type": idea.content_type,
        "topic": idea.topic,
        "examples": examples,
        "draft_tweet": llm.get("draft_tweet", ""),
        "thread_structure": llm.get("thread_structure", ""),
        "reference_notes": llm.get("reference_notes", ""),
        "remix_ideas": llm.get("remix_ideas", ""),
        "enrich_status": enrich_status,
        "enrich_reason": enrich_reason,
        "total_results": len(raw_results),
        "scored_results": len(scored),
        "queries": generate_twitter_queries(idea),
    }


# ---------------------------------------------------------------------------
# LinkedIn integration
# ---------------------------------------------------------------------------

LINKEDIN_TYPE_QUERY_MAP = {
    "Story": "entrepreneur story lessons learned",
    "BTS": "building a company behind the scenes",
    "Teach": "business leadership lessons",
    "Trend": "unpopular opinion entrepreneur",
    "Breakdown": "business strategy breakdown",
    "Reflection": "founder honest reflection journey",
    "Depth": "entrepreneur deep dive lessons"
}


def generate_linkedin_queries(idea: ContentIdea) -> List[str]:
    """Generate up to 3 LinkedIn search queries (professional phrases)."""
    queries = []

    def truncate_to_6_words(text: str) -> str:
        words = text.split()
        return ' '.join(words[:6]) if len(words) > 6 else text

    # 1. Format-specific professional phrase (highest value)
    queries.append(LINKEDIN_TYPE_QUERY_MAP.get(idea.content_type, "startup founder insights"))

    # 2. Topic as query
    if len(idea.topic.split()) <= 6:
        queries.append(idea.topic.lower())

    # 3. Best keyword combo
    all_kw = extract_keywords(idea.topic) + extract_keywords(idea.description)
    if len(all_kw) >= 2:
        queries.append(truncate_to_6_words(' '.join(all_kw[:4])))

    # Deduplicate
    seen = set()
    unique = []
    for q in queries:
        q_norm = q.lower().strip()
        if q_norm not in seen:
            seen.add(q_norm)
            unique.append(q_norm)

    if len(unique) < 2:
        unique.append("startup founder lessons")

    return unique[:3]


def fetch_linkedin_results(queries: List[str], token: Optional[str] = None) -> List[Dict[str, Any]]:
    """LinkedIn is LLM-only — no API calls. Always returns empty results."""
    return []


def normalize_linkedin_result(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Normalize Apify LinkedIn post result to standard schema."""
    post_id = raw.get("postId") or raw.get("id") or raw.get("urn")
    if not post_id:
        return None

    return {
        "post_id": str(post_id),
        "url": raw.get("postUrl", "") or raw.get("url", ""),
        "text": raw.get("text", "") or raw.get("commentary", ""),
        "author": raw.get("authorName", "") or raw.get("author", {}).get("name", "") if isinstance(raw.get("author"), dict) else raw.get("author", ""),
        "reactions": raw.get("totalReactionCount", 0) or raw.get("numLikes", 0),
        "comments": raw.get("commentsCount", 0) or raw.get("numComments", 0),
        "shares": raw.get("repostsCount", 0) or raw.get("numShares", 0),
        "created_at": raw.get("postedAtISO", None) or raw.get("postedAt", None) or raw.get("created_at", None),
    }


def score_linkedin_post(post: Dict[str, Any], keywords: List[str] = None) -> Dict[str, Any]:
    """Score a LinkedIn post: log10(reactions+1)*0.40 + log10(comments+1)*0.35 + log10(shares+1)*0.25 + boosts."""
    reactions = post.get("reactions", 0) or 0
    comments = post.get("comments", 0) or 0
    shares = post.get("shares", 0) or 0

    base = (math.log10(reactions + 1) * 0.40 +
            math.log10(comments + 1) * 0.35 +
            math.log10(shares + 1) * 0.25)

    recency_boost = 0.0
    created_at = post.get("created_at")
    if created_at:
        try:
            created_dt = date_parse(created_at)
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=datetime.timezone.utc)
            days_old = (datetime.datetime.now(datetime.timezone.utc) - created_dt).days
            if days_old <= 14:
                recency_boost = 0.2
            elif days_old <= 30:
                recency_boost = 0.1
        except Exception:
            pass

    relevance_boost = 0.0
    if keywords:
        text = post.get("text", "")
        if text:
            text_lower = text.lower()
            overlap = sum(1 for kw in keywords if kw.lower() in text_lower)
            relevance_boost = min(overlap * 0.05, 0.2)

    post["base_score"] = base
    post["recency_boost"] = recency_boost
    post["relevance_boost"] = relevance_boost
    post["final_score"] = base + recency_boost + relevance_boost
    return post


def process_linkedin_results(raw_results: List[Dict[str, Any]], keywords: List[str] = None) -> List[Dict[str, Any]]:
    """Normalize → dedup → filter old → score → sort LinkedIn posts."""
    normalized = [normalize_linkedin_result(r) for r in raw_results]
    normalized = [n for n in normalized if n is not None]

    seen = set()
    deduped = []
    for p in normalized:
        pid = p.get("post_id")
        if pid and pid not in seen:
            seen.add(pid)
            deduped.append(p)

    filtered = filter_old_results(deduped, max_age_days=120)
    scored = [score_linkedin_post(p, keywords) for p in filtered]
    return sorted(scored, key=lambda x: x.get("final_score", 0), reverse=True)


def select_top_linkedin_posts(scored: List[Dict[str, Any]], count: int = 3) -> List[Dict[str, Any]]:
    """Select top N LinkedIn posts as examples."""
    if not scored:
        return []
    examples = []
    for post in scored[:count]:
        examples.append({
            "url": post["url"],
            "text": post["text"],
            "author": post["author"],
            "reactions": post["reactions"],
            "comments": post["comments"],
            "shares": post["shares"],
            "created_at": post["created_at"],
            "final_score": post["final_score"],
        })
    return examples


def generate_linkedin_content(idea: ContentIdea, examples: List[Dict[str, Any]], client=None) -> Dict[str, Any]:
    """Generate LinkedIn content via LLM: draft_post, reference_notes, remix_ideas."""
    empty = {"draft_post": "", "reference_notes": "", "remix_ideas": ""}

    if client is None:
        if not os.environ.get("OPENAI_API_KEY"):
            return empty
        client = openai.OpenAI()

    system_message = (
        "You are a LinkedIn content strategist for Logara, a startup documenting its founder journey. "
        "You craft high-performing professional posts optimized for the LinkedIn algorithm. "
        "Respond in valid JSON only."
    )

    examples_text = ""
    for i, ex in enumerate(examples, 1):
        examples_text += f"\nExample {i}:\n"
        examples_text += f"  - Text: {ex.get('text', 'N/A')[:500]}\n"
        examples_text += f"  - Reactions: {ex.get('reactions', 0):,} | Comments: {ex.get('comments', 0):,}\n"
        examples_text += f"  - Author: {ex.get('author', 'unknown')}\n"

    user_message = f"""Content Idea:
- Type: {idea.content_type}
- Topic: {idea.topic}
- Description: {idea.description}

Top Example LinkedIn Posts:{examples_text}

Brand Context:
- Logara is a startup documenting its founder journey
- Tone: Professional but authentic, insightful, conversational

LinkedIn Algorithm Best Practices (YOU MUST follow these):
- HOOK (first 2-3 lines before "see more"): Use a contrarian take, a surprising stat, or start mid-story to create a curiosity gap. This is the most important part — if the hook fails, nobody reads the rest.
- FORMAT: Short paragraphs (1-2 sentences max). Use line breaks liberally. Write at 5th-8th grade reading level. No walls of text.
- DWELL TIME: Structure the post so readers spend time on it — use lists, bold key phrases, or a narrative arc that builds.
- CTA: End with a genuine question that invites thoughtful comments (5+ word comments are worth 2x a like to the algorithm). Never use engagement bait like "Like if you agree".
- LENGTH: 150-300 words. Enough depth to be valuable, short enough to hold attention.
- NO emojis overload, no hashtag spam. Max 3-5 hashtags at the end.

Instructions:
1. draft_post: A full LinkedIn post following the best practices above. Must have: a scroll-stopping hook, short-paragraph body with real insight, and a comment-driving question at the end.
2. reference_notes: For each example, one sentence on why it works and what to learn
3. remix_ideas: 2-3 post angle ideas as one-liner bullets

Return JSON:
{{
  "draft_post": "the full linkedin post text",
  "reference_notes": ["note for ex1", "note for ex2", "note for ex3"],
  "remix_ideas": ["angle 1", "angle 2", "angle 3"]
}}"""

    def make_api_call():
        return client.chat.completions.create(
            model="gpt-4o-mini",
            response_format={"type": "json_object"},
            temperature=0.8,
            max_tokens=1000,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ]
        )

    try:
        response = _call_with_retry(make_api_call)
        if not response or response == []:
            return empty

        parsed = json.loads(response.choices[0].message.content)
        ref_notes = parsed.get("reference_notes", [])
        remix_list = parsed.get("remix_ideas", [])

        return {
            "draft_post": parsed.get("draft_post", ""),
            "reference_notes": "\n".join(f"- {n}" for n in ref_notes) if ref_notes else "",
            "remix_ideas": "\n".join(f"- {r}" for r in remix_list) if remix_list else "",
        }
    except (json.JSONDecodeError, KeyError, IndexError, AttributeError) as e:
        print(f"Error parsing LinkedIn LLM response: {e}", file=sys.stderr)
        return empty


def enrich_row_linkedin(idea: ContentIdea, raw_results: List[Dict[str, Any]], openai_client=None) -> Dict[str, Any]:
    """Orchestrate full LinkedIn enrichment pipeline for a single row."""
    keywords = extract_keywords(idea.topic) + extract_keywords(idea.description)
    scored = process_linkedin_results(raw_results, keywords=keywords)
    examples = select_top_linkedin_posts(scored)

    enrich_status = "ok"
    enrich_reason = ""

    if len(scored) == 0:
        # No API results — generate content from idea alone (LLM-only fallback)
        llm = generate_linkedin_content(idea, [], client=openai_client)
        llm_has_content = any(llm.get(k, "") for k in ["draft_post", "remix_ideas"])
        if llm_has_content:
            enrich_status = "partial"
            enrich_reason = "LLM-only (no API results)"
        else:
            enrich_status = "skipped"
            enrich_reason = "No API results and LLM generation failed"
    else:
        llm = generate_linkedin_content(idea, examples, client=openai_client)
        llm_failed = all(llm.get(k, "") == "" for k in ["draft_post", "reference_notes", "remix_ideas"])
        reasons = []
        if llm_failed:
            reasons.append("LLM generation failed")
        if len(examples) < 2:
            reasons.append(f"Only {len(examples)} example(s), need >=2")
        if reasons:
            enrich_status = "partial"
            enrich_reason = "; ".join(reasons)

    return {
        "row_number": idea.row_number,
        "content_type": idea.content_type,
        "topic": idea.topic,
        "examples": examples,
        "draft_post": llm.get("draft_post", ""),
        "reference_notes": llm.get("reference_notes", ""),
        "remix_ideas": llm.get("remix_ideas", ""),
        "enrich_status": enrich_status,
        "enrich_reason": enrich_reason,
        "total_results": len(raw_results),
        "scored_results": len(scored),
        "queries": generate_linkedin_queries(idea),
    }


# ---------------------------------------------------------------------------
# TikTok LLM content generation (original)
# ---------------------------------------------------------------------------

def generate_llm_content(idea: ContentIdea, examples: List[Dict[str, Any]], audio: Dict[str, Any], client=None) -> Dict[str, Any]:
    """
    Generate LLM content for enrichment: audio_fit_reason, hook_summaries, remix_ideas.

    Uses OpenAI GPT-4o-mini to generate actionable creative text based on:
    - Content idea (type, topic, description)
    - Top example videos (caption, engagement, audio)
    - Selected audio track

    Args:
        idea: ContentIdea to generate content for
        examples: List of example video dicts from select_top_examples()
        audio: Audio dict from select_audio()
        client: Optional OpenAI client for testing (creates one if None)

    Returns:
        Dict with audio_fit_reason, ex1_hook_summary, ex2_hook_summary,
        ex3_hook_summary, remix_ideas. Returns empty strings on API failure.
    """
    # Step 1: Client setup
    if client is None:
        # Check if OPENAI_API_KEY is set
        if not os.environ.get("OPENAI_API_KEY"):
            # No API key - return empty fields
            return {
                "audio_fit_reason": "",
                "ex1_hook_summary": "",
                "ex2_hook_summary": "",
                "ex3_hook_summary": "",
                "remix_ideas": "",
                "hook_scripts": ""
            }
        client = openai.OpenAI()  # Reads OPENAI_API_KEY from env automatically

    # Step 2: Build prompt with brand context and type-specific direction
    system_message = (
        "You are a hype coach content strategist for Logara, a startup documenting its founder journey. "
        "You give energetic, direct, filmable content direction. Respond in valid JSON only."
    )

    # Build examples section
    examples_text = ""
    for i, ex in enumerate(examples, 1):
        examples_text += f"\nExample {i}:\n"
        examples_text += f"  - Caption: {ex.get('caption', 'N/A')}\n"
        examples_text += f"  - Views: {ex.get('views', 0):,} | Likes: {ex.get('likes', 0):,}\n"
        examples_text += f"  - Author: @{ex.get('author_username', 'unknown')}\n"
        examples_text += f"  - Audio: {ex.get('audio_title', 'N/A')}\n"

    # Adjust tone based on audio confidence
    audio_confidence = audio.get('audio_confidence', 'low')
    if audio_confidence == 'high':
        audio_instruction = "Use assertive language for audio_fit_reason (e.g., 'Use this track - ...')"
    else:
        audio_instruction = "Use suggestive language for audio_fit_reason (e.g., 'Consider this audio, but feel free to pick your own - ...')"

    # Type-specific framing
    type_framing = {
        "Teach": "Educational angles - focus on what viewers will learn",
        "Story": "Narrative arcs - emphasize storytelling structure",
        "BTS": "Behind-the-scenes framing - show the process/reality",
        "Trend": "Trending formats - leverage popular patterns",
        "Breakdown": "Analysis angles - break down concepts clearly",
        "Reflection": "Personal insight - authentic founder perspective",
        "Depth": "Longform depth - substantive exploration"
    }
    type_hint = type_framing.get(idea.content_type, "Authentic founder content")

    user_message = f"""Content Idea:
- Type: {idea.content_type}
- Topic: {idea.topic}
- Description: {idea.description}

Audio Track:
- Title: {audio.get('audio_title', 'N/A')}
- Author: {audio.get('audio_author', 'N/A')}
- Confidence: {audio_confidence}

Top Examples:{examples_text}

Brand Context:
- Logara is a startup documenting its founder journey
- Tone: Hopecore-light (optimistic, authentic, not forced)
- Type framing: {type_hint}

Instructions:
1. audio_fit_reason: One sentence about the audio's GENERAL vibe and usage potential (not tied to this specific idea). {audio_instruction}
2. hook_summaries: For each example, write 2 sentences: (a) what the video did (reference the caption), (b) why it works/what to learn
3. remix_ideas: 2-3 filmable one-liner bullets tailored to {idea.content_type} format. Concrete actions, not abstract advice.
4. hook_scripts: 2-3 short hook opening lines (first 3 seconds of a video) tailored to this content idea. Each hook should be a single sentence the founder can say to camera.

Return JSON:
{{
  "audio_fit_reason": "one sentence",
  "hook_summaries": ["2 sentences for ex1", "2 sentences for ex2", "2 sentences for ex3"],
  "remix_ideas": ["bullet 1", "bullet 2", "bullet 3"],
  "hook_scripts": ["hook line 1", "hook line 2", "hook line 3"]
}}"""

    # Step 3: Define API call function for retry wrapper
    def make_api_call():
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            response_format={"type": "json_object"},
            temperature=0.8,
            max_tokens=700,
            messages=[
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ]
        )
        return response

    # Step 4: Call API with retry wrapper
    try:
        response = _call_with_retry(make_api_call)

        # If retry failed, response will be []
        if not response or response == []:
            return {
                "audio_fit_reason": "",
                "ex1_hook_summary": "",
                "ex2_hook_summary": "",
                "ex3_hook_summary": "",
                "remix_ideas": "",
                "hook_scripts": ""
            }

        # Step 5: Parse response
        content = response.choices[0].message.content
        parsed = json.loads(content)

        # Extract fields with defaults
        audio_fit_reason = parsed.get("audio_fit_reason", "")
        hook_summaries = parsed.get("hook_summaries", [])
        remix_ideas_list = parsed.get("remix_ideas", [])
        hook_scripts_list = parsed.get("hook_scripts", [])

        # Format remix_ideas as bulleted string
        if remix_ideas_list:
            remix_ideas = "\n".join(f"- {idea}" for idea in remix_ideas_list)
        else:
            remix_ideas = ""

        # Format hook_scripts as bulleted string
        if hook_scripts_list:
            hook_scripts = "\n".join(f"- {h}" for h in hook_scripts_list)
        else:
            hook_scripts = ""

        # Step 6: Return structured dict
        return {
            "audio_fit_reason": audio_fit_reason,
            "ex1_hook_summary": hook_summaries[0] if len(hook_summaries) > 0 else "",
            "ex2_hook_summary": hook_summaries[1] if len(hook_summaries) > 1 else "",
            "ex3_hook_summary": hook_summaries[2] if len(hook_summaries) > 2 else "",
            "remix_ideas": remix_ideas,
            "hook_scripts": hook_scripts
        }

    except (json.JSONDecodeError, KeyError, IndexError, AttributeError) as e:
        print(f"Error parsing LLM response: {e}", file=sys.stderr)
        return {
            "audio_fit_reason": "",
            "ex1_hook_summary": "",
            "ex2_hook_summary": "",
            "ex3_hook_summary": "",
            "remix_ideas": "",
            "hook_scripts": ""
        }


def enrich_row(idea: ContentIdea, raw_results: List[Dict[str, Any]], openai_client=None) -> Dict[str, Any]:
    """
    Orchestrate full Phase 3 pipeline for a single content row.

    This is the main Phase 3 entry point that chains together:
    1. Keyword extraction for relevance scoring
    2. Result processing (normalize, dedup, filter, score, sort)
    3. Top example selection
    4. Audio selection

    Args:
        idea: ContentIdea to process
        raw_results: List of raw Apify result dicts from TikTok search

    Returns:
        Enrichment dict containing row metadata, top 3 examples, selected audio,
        result counts, and queries used.
    """
    # Step 1: Extract keywords for relevance scoring
    topic_keywords = extract_keywords(idea.topic)
    desc_keywords = extract_keywords(idea.description)
    keywords = topic_keywords + desc_keywords

    # Step 2: Process results (normalize, dedup, filter, score, sort)
    scored = process_results(raw_results, keywords=keywords)

    # Step 3: Select top 3 examples
    examples = select_top_examples(scored)

    # Step 4: Select audio
    audio = select_audio(scored, examples)

    # Step 5: Determine enrich_status per LOG-01
    enrich_status = "ok"
    enrich_reason = ""

    if len(scored) == 0:
        # No results at all - skip LLM call
        llm_content = {
            "audio_fit_reason": "",
            "ex1_hook_summary": "",
            "ex2_hook_summary": "",
            "ex3_hook_summary": "",
            "remix_ideas": "",
            "hook_scripts": ""
        }
        enrich_status = "skipped"
        enrich_reason = "No scored results"
    else:
        # Generate LLM content
        llm_content = generate_llm_content(idea, examples, audio, client=openai_client)

        # Check if LLM generation failed (all fields empty)
        llm_failed = all(
            llm_content.get(key, "") == ""
            for key in ["audio_fit_reason", "ex1_hook_summary", "ex2_hook_summary",
                       "ex3_hook_summary", "remix_ideas", "hook_scripts"]
        )

        # Determine status based on LOG-01 criteria
        has_audio = bool(audio.get("audio_title"))
        has_enough_examples = len(examples) >= 2
        reasons = []

        if llm_failed:
            reasons.append("LLM generation failed")
        if not has_enough_examples:
            reasons.append(f"Only {len(examples)} example(s), need >=2")
        if not has_audio:
            reasons.append("No audio selected")

        if reasons:
            enrich_status = "partial"
            enrich_reason = "; ".join(reasons)

    # Step 6: Build and return enrichment dict with LLM fields
    return {
        "row_number": idea.row_number,
        "content_type": idea.content_type,
        "topic": idea.topic,
        "topic_keywords": topic_keywords,
        "examples": examples,
        "audio": audio,
        "audio_fit_reason": llm_content.get("audio_fit_reason", ""),
        "ex1_hook_summary": llm_content.get("ex1_hook_summary", ""),
        "ex2_hook_summary": llm_content.get("ex2_hook_summary", ""),
        "ex3_hook_summary": llm_content.get("ex3_hook_summary", ""),
        "ex1_audio_title": examples[0]["audio_title"] if len(examples) > 0 else "",
        "ex2_audio_title": examples[1]["audio_title"] if len(examples) > 1 else "",
        "ex3_audio_title": examples[2]["audio_title"] if len(examples) > 2 else "",
        "remix_ideas": llm_content.get("remix_ideas", ""),
        "hook_scripts": llm_content.get("hook_scripts", ""),
        "enrich_status": enrich_status,
        "enrich_reason": enrich_reason,
        "total_results": len(raw_results),
        "scored_results": len(scored),
        "queries": generate_queries(idea)
    }


def _style_sheet(ws, wrap_columns: List[str] = None) -> None:
    """Apply shared styling to a worksheet: bold headers, freeze panes, auto-width."""
    headers = [cell.value for cell in ws[1]]

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Wrap text on specified columns
    if wrap_columns:
        for col_name in wrap_columns:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in column_cells:
            if cell.value:
                content_length = min(len(str(cell.value)), 50)
                max_length = max(max_length, content_length)
        ws.column_dimensions[col_letter].width = max_length + 2


def write_enriched_excel(ideas: List[ContentIdea], enrichments: List[Dict[str, Any]],
                         output_path: str,
                         twitter_enrichments: List[Dict[str, Any]] = None,
                         linkedin_enrichments: List[Dict[str, Any]] = None) -> None:
    """
    Write enriched content to Excel with 3 sheets: TikTok, Twitter, LinkedIn.

    Args:
        ideas: List of original ContentIdea objects
        enrichments: List of TikTok enrichment dicts from enrich_row()
        output_path: Path to write output Excel file
        twitter_enrichments: List of Twitter enrichment dicts from enrich_row_twitter()
        linkedin_enrichments: List of LinkedIn enrichment dicts from enrich_row_linkedin()
    """
    wb = Workbook()

    # ---- Sheet 1: TikTok ----
    ws_tiktok = wb.active
    ws_tiktok.title = "TikTok"

    tiktok_headers = [
        "Day", "Date", "Type", "Topic", "Description",
        "Topic Keywords", "Search Queries",
        "Ex1 URL", "Ex1 Views", "Ex1 Likes", "Ex1 Comments", "Ex1 Shares",
        "Ex1 Author", "Ex1 Caption", "Ex1 Audio", "Ex1 Hook Summary",
        "Ex2 URL", "Ex2 Views", "Ex2 Likes", "Ex2 Comments", "Ex2 Shares",
        "Ex2 Author", "Ex2 Caption", "Ex2 Audio", "Ex2 Hook Summary",
        "Ex3 URL", "Ex3 Views", "Ex3 Likes", "Ex3 Comments", "Ex3 Shares",
        "Ex3 Author", "Ex3 Caption", "Ex3 Audio", "Ex3 Hook Summary",
        "Audio Title", "Audio Author", "Audio Confidence", "Audio URL", "Audio Fit Reason",
        "Remix Ideas", "Hook Scripts",
        "Enrich Status", "Enrich Reason"
    ]
    ws_tiktok.append(tiktok_headers)

    enrichment_map = {e["row_number"]: e for e in enrichments}
    for idea in ideas:
        e = enrichment_map.get(idea.row_number, {})
        examples = e.get("examples", [])
        ex1 = examples[0] if len(examples) > 0 else {}
        ex2 = examples[1] if len(examples) > 1 else {}
        ex3 = examples[2] if len(examples) > 2 else {}
        audio = e.get("audio", {})

        ws_tiktok.append([
            idea.date.strftime("%A"), idea.date.strftime("%Y-%m-%d"),
            idea.content_type, idea.topic, idea.description,
            ", ".join(e.get("topic_keywords", [])),
            " | ".join(e.get("queries", [])),
            ex1.get("url", ""), ex1.get("views", ""), ex1.get("likes", ""),
            ex1.get("comments", ""), ex1.get("shares", ""),
            ex1.get("author_username", ""), ex1.get("caption", ""),
            ex1.get("audio_title", ""), e.get("ex1_hook_summary", ""),
            ex2.get("url", ""), ex2.get("views", ""), ex2.get("likes", ""),
            ex2.get("comments", ""), ex2.get("shares", ""),
            ex2.get("author_username", ""), ex2.get("caption", ""),
            ex2.get("audio_title", ""), e.get("ex2_hook_summary", ""),
            ex3.get("url", ""), ex3.get("views", ""), ex3.get("likes", ""),
            ex3.get("comments", ""), ex3.get("shares", ""),
            ex3.get("author_username", ""), ex3.get("caption", ""),
            ex3.get("audio_title", ""), e.get("ex3_hook_summary", ""),
            audio.get("audio_title", ""), audio.get("audio_author", ""),
            audio.get("audio_confidence", ""), audio.get("audio_url", ""),
            e.get("audio_fit_reason", ""),
            e.get("remix_ideas", ""), e.get("hook_scripts", ""),
            e.get("enrich_status", ""), e.get("enrich_reason", ""),
        ])

    _style_sheet(ws_tiktok, wrap_columns=["Remix Ideas", "Hook Scripts"])

    # ---- Sheet 2: Twitter ----
    ws_twitter = wb.create_sheet("Twitter")

    twitter_headers = [
        "Day", "Date", "Type", "Topic", "Description", "Search Queries",
        "Draft Tweet", "Thread Structure",
        "Ex1 URL", "Ex1 Text", "Ex1 Likes", "Ex1 Retweets", "Ex1 Author",
        "Ex2 URL", "Ex2 Text", "Ex2 Likes", "Ex2 Retweets", "Ex2 Author",
        "Ex3 URL", "Ex3 Text", "Ex3 Likes", "Ex3 Retweets", "Ex3 Author",
        "Reference Notes", "Remix Ideas",
        "Enrich Status", "Enrich Reason"
    ]
    ws_twitter.append(twitter_headers)

    tw_map = {e["row_number"]: e for e in (twitter_enrichments or [])}
    for idea in ideas:
        te = tw_map.get(idea.row_number, {})
        examples = te.get("examples", [])
        ex1 = examples[0] if len(examples) > 0 else {}
        ex2 = examples[1] if len(examples) > 1 else {}
        ex3 = examples[2] if len(examples) > 2 else {}

        ws_twitter.append([
            idea.date.strftime("%A"), idea.date.strftime("%Y-%m-%d"),
            idea.content_type, idea.topic, idea.description,
            " | ".join(te.get("queries", [])),
            te.get("draft_tweet", ""), te.get("thread_structure", ""),
            ex1.get("url", ""), ex1.get("text", ""), ex1.get("likes", ""),
            ex1.get("retweets", ""), ex1.get("author", ""),
            ex2.get("url", ""), ex2.get("text", ""), ex2.get("likes", ""),
            ex2.get("retweets", ""), ex2.get("author", ""),
            ex3.get("url", ""), ex3.get("text", ""), ex3.get("likes", ""),
            ex3.get("retweets", ""), ex3.get("author", ""),
            te.get("reference_notes", ""), te.get("remix_ideas", ""),
            te.get("enrich_status", ""), te.get("enrich_reason", ""),
        ])

    _style_sheet(ws_twitter, wrap_columns=["Draft Tweet", "Thread Structure", "Reference Notes", "Remix Ideas"])

    # ---- Sheet 3: LinkedIn ----
    ws_linkedin = wb.create_sheet("LinkedIn")

    linkedin_headers = [
        "Day", "Date", "Type", "Topic", "Description", "Search Queries",
        "Draft Post",
        "Ex1 URL", "Ex1 Text", "Ex1 Reactions", "Ex1 Comments", "Ex1 Author",
        "Ex2 URL", "Ex2 Text", "Ex2 Reactions", "Ex2 Comments", "Ex2 Author",
        "Ex3 URL", "Ex3 Text", "Ex3 Reactions", "Ex3 Comments", "Ex3 Author",
        "Reference Notes", "Remix Ideas",
        "Enrich Status", "Enrich Reason"
    ]
    ws_linkedin.append(linkedin_headers)

    li_map = {e["row_number"]: e for e in (linkedin_enrichments or [])}
    for idea in ideas:
        le = li_map.get(idea.row_number, {})
        examples = le.get("examples", [])
        ex1 = examples[0] if len(examples) > 0 else {}
        ex2 = examples[1] if len(examples) > 1 else {}
        ex3 = examples[2] if len(examples) > 2 else {}

        ws_linkedin.append([
            idea.date.strftime("%A"), idea.date.strftime("%Y-%m-%d"),
            idea.content_type, idea.topic, idea.description,
            " | ".join(le.get("queries", [])),
            le.get("draft_post", ""),
            ex1.get("url", ""), ex1.get("text", ""), ex1.get("reactions", ""),
            ex1.get("comments", ""), ex1.get("author", ""),
            ex2.get("url", ""), ex2.get("text", ""), ex2.get("reactions", ""),
            ex2.get("comments", ""), ex2.get("author", ""),
            ex3.get("url", ""), ex3.get("text", ""), ex3.get("reactions", ""),
            ex3.get("comments", ""), ex3.get("author", ""),
            le.get("reference_notes", ""), le.get("remix_ideas", ""),
            le.get("enrich_status", ""), le.get("enrich_reason", ""),
        ])

    _style_sheet(ws_linkedin, wrap_columns=["Draft Post", "Reference Notes", "Remix Ideas"])

    wb.save(output_path)


def build_run_log(ideas: List[ContentIdea], enrichments: List[Dict[str, Any]],
                  input_path: str, output_path: str, start_time: float, end_time: float,
                  twitter_enrichments: List[Dict[str, Any]] = None,
                  linkedin_enrichments: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """
    Build complete run log with summary and per-row diagnostics for all platforms.

    Args:
        ideas: List of original ContentIdea objects
        enrichments: List of TikTok enrichment dicts from enrich_row()
        input_path: Path to input Excel file
        output_path: Path to output Excel file
        start_time: time.time() at run start
        end_time: time.time() at run end
        twitter_enrichments: List of Twitter enrichment dicts
        linkedin_enrichments: List of LinkedIn enrichment dicts

    Returns:
        Dict with run_summary and rows array (each row has tiktok/twitter/linkedin sections)
    """
    timestamp = datetime.datetime.now(datetime.timezone.utc).isoformat()
    duration_seconds = round(end_time - start_time, 1)

    # Count status values per platform
    def count_statuses(enrich_list):
        counts = {"ok": 0, "partial": 0, "skipped": 0, "error": 0}
        for e in enrich_list:
            s = e.get("enrich_status", "error")
            counts[s] = counts.get(s, 0) + 1
        return counts

    tiktok_status = count_statuses(enrichments)
    twitter_status = count_statuses(twitter_enrichments or [])
    linkedin_status = count_statuses(linkedin_enrichments or [])

    run_summary = {
        "timestamp": timestamp,
        "input_file": os.path.basename(input_path),
        "output_file": os.path.basename(output_path),
        "total_rows": len(ideas),
        "tiktok_status_counts": tiktok_status,
        "twitter_status_counts": twitter_status,
        "linkedin_status_counts": linkedin_status,
        "duration_seconds": duration_seconds
    }

    # Build rows array
    rows = []
    tiktok_map = {e["row_number"]: e for e in enrichments}
    tw_map = {e["row_number"]: e for e in (twitter_enrichments or [])}
    li_map = {e["row_number"]: e for e in (linkedin_enrichments or [])}

    for idea in ideas:
        tk = tiktok_map.get(idea.row_number, {})
        tw = tw_map.get(idea.row_number, {})
        li = li_map.get(idea.row_number, {})

        audio = tk.get("audio", {})
        tk_examples = tk.get("examples", [])

        row_data = {
            "row_number": idea.row_number,
            "content_type": idea.content_type,
            "topic": idea.topic,
            "tiktok": {
                "queries": tk.get("queries", []),
                "total_results": tk.get("total_results", 0),
                "scored_results": tk.get("scored_results", 0),
                "chosen_audio": {"title": audio.get("audio_title", ""), "confidence": audio.get("audio_confidence", "")},
                "example_urls": [ex.get("url", "") for ex in tk_examples[:3]],
                "enrich_status": tk.get("enrich_status", "error"),
                "enrich_reason": tk.get("enrich_reason", ""),
            },
            "twitter": {
                "queries": tw.get("queries", []),
                "total_results": tw.get("total_results", 0),
                "scored_results": tw.get("scored_results", 0),
                "example_urls": [ex.get("url", "") for ex in tw.get("examples", [])[:3]],
                "enrich_status": tw.get("enrich_status", "error"),
                "enrich_reason": tw.get("enrich_reason", ""),
            },
            "linkedin": {
                "queries": li.get("queries", []),
                "total_results": li.get("total_results", 0),
                "scored_results": li.get("scored_results", 0),
                "example_urls": [ex.get("url", "") for ex in li.get("examples", [])[:3]],
                "enrich_status": li.get("enrich_status", "error"),
                "enrich_reason": li.get("enrich_reason", ""),
            },
        }
        rows.append(row_data)

    return {
        "run_summary": run_summary,
        "rows": rows
    }


def save_run_log(run_log: Dict[str, Any], log_path: str) -> None:
    """
    Save run log dict to JSON file.

    Args:
        run_log: Run log dict from build_run_log()
        log_path: Path to write JSON file
    """
    with open(log_path, 'w', encoding='utf-8') as f:
        json.dump(run_log, f, indent=2, default=str)


def main():
    """
    Main CLI entry point for content calendar enrichment.

    Enriches each content idea across TikTok, Twitter, and LinkedIn.
    Output: Excel with 3 sheets + run_log.json.
    """
    # Step 1: Parse args
    parser = argparse.ArgumentParser(
        description="Enrich a content calendar with TikTok, Twitter, and LinkedIn trends."
    )
    parser.add_argument("--input", default="Content ideas.xlsx",
        help="Path to input Excel file (default: Content ideas.xlsx)")
    parser.add_argument("--output", default=None,
        help="Path to output Excel file (default: Enriched <input>_YYYY-MM-DD.xlsx)")
    parser.add_argument("--dry-run", action="store_true",
        help="Validate input and show what would be processed without making API calls")
    parser.add_argument("--max-rows", type=int, default=None,
        help="Limit processing to first N rows (useful for testing)")
    parser.add_argument("--estimate", action="store_true",
        help="Show projected Apify cost and confirm before running")

    args = parser.parse_args()

    # Step 2: Validate input file exists
    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    # Step 3: Environment variable check (skip for --dry-run and --estimate)
    if not args.dry_run and not args.estimate:
        missing_vars = []
        if not os.environ.get("RAPIDAPI_KEY"):
            missing_vars.append("RAPIDAPI_KEY")
        if not os.environ.get("OPENAI_API_KEY"):
            missing_vars.append("OPENAI_API_KEY")

        if missing_vars:
            for var in missing_vars:
                print(f"Missing environment variable: {var}", file=sys.stderr)
            sys.exit(1)

    # Step 4: Load content ideas
    ideas, skipped = load_content_ideas(args.input)
    print(f"Loaded {len(ideas)} content ideas from {args.input}")
    if skipped:
        print(f"Skipped {len(skipped)} rows with missing fields")

    # Step 4b: Apply --max-rows limit
    if args.max_rows is not None and args.max_rows > 0:
        ideas = ideas[:args.max_rows]
        print(f"Limited to first {len(ideas)} rows (--max-rows {args.max_rows})")

    # Step 5: Handle --dry-run
    if args.dry_run:
        print(f"\nPlatforms: TikTok, Twitter, LinkedIn")
        for idea in ideas:
            print(f"  Row {idea.row_number}: {idea.content_type} | {idea.topic}")
        print(f"\nDry run complete. {len(ideas)} rows x 3 platforms would be processed.")
        if skipped:
            for skip_info in skipped:
                print(f"  Would skip row {skip_info['row_number']}: missing {skip_info['field']}")
        sys.exit(0)

    # Step 5b: Handle --estimate
    if args.estimate:
        api_rows = [idea for idea in ideas if has_enough_context(idea)]
        llm_only_rows = len(ideas) - len(api_rows)
        tk_queries = len(api_rows) * 3
        tw_queries = len(api_rows) * 3
        tk_requests = tk_queries  # 1 RapidAPI request per query
        tw_requests = tw_queries  # 1 RapidAPI request per query

        print(f"\nEstimated API usage for {len(ideas)} rows:")
        print(f"  API rows:      {len(api_rows)} (LLM-only: {llm_only_rows})")
        print(f"  TikTok:   {tk_queries} queries = {tk_requests} RapidAPI requests (free tier)")
        print(f"  Twitter:  {tw_queries} queries = {tw_requests} RapidAPI requests (free: 1000/month)")
        print(f"  LinkedIn: LLM-only (no API calls)")
        print(f"  Total RapidAPI requests: {tk_requests + tw_requests}")
        print(f"  Cost: $0.00 (RapidAPI free tier)")

        confirm = input("\nProceed? [y/N] ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            sys.exit(0)

    # Step 6: Process rows across all platforms
    start_time = time.time()
    tiktok_enrichments = []
    twitter_enrichments = []
    linkedin_enrichments = []
    openai_client = openai.OpenAI()

    for i, idea in enumerate(ideas, 1):
        print(f"\nRow {i}/{len(ideas)}: {idea.content_type} | {idea.topic}")
        use_api = has_enough_context(idea)
        if not use_api:
            print(f"  Low context — LLM-only mode (saving API costs)")

        # --- TikTok ---
        try:
            tiktok_queries = generate_queries(idea)
            tiktok_raw = fetch_tiktok_results(tiktok_queries) if use_api else []
            tiktok_enriched = enrich_row(idea, tiktok_raw, openai_client=openai_client)
            tiktok_enrichments.append(tiktok_enriched)
            print(f"  TikTok: {tiktok_enriched.get('enrich_status', 'ok')}")
        except Exception as e:
            print(f"  TikTok: error - {e}", file=sys.stderr)
            tiktok_enrichments.append({
                "row_number": idea.row_number, "content_type": idea.content_type,
                "topic": idea.topic, "topic_keywords": [], "examples": [],
                "audio": {"audio_title": "", "audio_author": "", "audio_id": "", "audio_url": "", "audio_confidence": "low"},
                "audio_fit_reason": "",
                "ex1_hook_summary": "", "ex2_hook_summary": "", "ex3_hook_summary": "",
                "ex1_audio_title": "", "ex2_audio_title": "", "ex3_audio_title": "",
                "remix_ideas": "", "hook_scripts": "",
                "enrich_status": "error", "enrich_reason": str(e),
                "total_results": 0, "scored_results": 0, "queries": [],
            })

        # --- Twitter ---
        try:
            twitter_queries = generate_twitter_queries(idea)
            twitter_raw = fetch_twitter_results(twitter_queries) if use_api else []
            twitter_enriched = enrich_row_twitter(idea, twitter_raw, openai_client=openai_client)
            twitter_enrichments.append(twitter_enriched)
            print(f"  Twitter: {twitter_enriched.get('enrich_status', 'ok')}")
        except Exception as e:
            print(f"  Twitter: error - {e}", file=sys.stderr)
            twitter_enrichments.append({
                "row_number": idea.row_number, "content_type": idea.content_type,
                "topic": idea.topic, "examples": [],
                "draft_tweet": "", "thread_structure": "",
                "reference_notes": "", "remix_ideas": "",
                "enrich_status": "error", "enrich_reason": str(e),
                "total_results": 0, "scored_results": 0, "queries": [],
            })

        # --- LinkedIn (LLM-only, no API) ---
        try:
            linkedin_queries = generate_linkedin_queries(idea)
            linkedin_enriched = enrich_row_linkedin(idea, [], openai_client=openai_client)
            linkedin_enrichments.append(linkedin_enriched)
            print(f"  LinkedIn: {linkedin_enriched.get('enrich_status', 'ok')}")
        except Exception as e:
            print(f"  LinkedIn: error - {e}", file=sys.stderr)
            linkedin_enrichments.append({
                "row_number": idea.row_number, "content_type": idea.content_type,
                "topic": idea.topic, "examples": [],
                "draft_post": "", "reference_notes": "", "remix_ideas": "",
                "enrich_status": "error", "enrich_reason": str(e),
                "total_results": 0, "scored_results": 0, "queries": [],
            })

    end_time = time.time()

    # Step 7: Compute output path
    if args.output is None:
        input_dir = os.path.dirname(args.input) or "."
        input_basename = os.path.basename(args.input)
        input_name, input_ext = os.path.splitext(input_basename)
        now_str = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        output_filename = f"Enriched {input_name}_{now_str}.xlsx"
        output_path = os.path.join(input_dir, output_filename)
    else:
        output_path = args.output

    # Step 8: Write outputs
    write_enriched_excel(ideas, tiktok_enrichments, output_path,
                         twitter_enrichments=twitter_enrichments,
                         linkedin_enrichments=linkedin_enrichments)
    print(f"\nEnriched Excel written to: {output_path}")

    # Build and save run log
    run_log = build_run_log(ideas, tiktok_enrichments, args.input, output_path,
                            start_time, end_time,
                            twitter_enrichments=twitter_enrichments,
                            linkedin_enrichments=linkedin_enrichments)
    log_dir = os.path.dirname(output_path) or "."
    log_filename = f"run_log_{datetime.datetime.now().strftime('%Y-%m-%d_%H%M%S')}.json"
    log_path = os.path.join(log_dir, log_filename)
    save_run_log(run_log, log_path)
    print(f"Run log written to: {log_path}")

    # Print summary
    summary = run_log["run_summary"]
    tk_sc = summary["tiktok_status_counts"]
    tw_sc = summary["twitter_status_counts"]
    li_sc = summary["linkedin_status_counts"]
    print(f"\nDone! {summary['total_rows']} rows x 3 platforms in {summary['duration_seconds']}s")
    print(f"  TikTok  - ok: {tk_sc['ok']} | partial: {tk_sc['partial']} | skipped: {tk_sc['skipped']} | error: {tk_sc['error']}")
    print(f"  Twitter - ok: {tw_sc['ok']} | partial: {tw_sc['partial']} | skipped: {tw_sc['skipped']} | error: {tw_sc['error']}")
    print(f"  LinkedIn- ok: {li_sc['ok']} | partial: {li_sc['partial']} | skipped: {li_sc['skipped']} | error: {li_sc['error']}")


if __name__ == "__main__":
    main()
